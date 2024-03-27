import openpyxl
import webbrowser

wb = openpyxl.load_workbook("ZAVIYE_DATABASE_v1.0.xlsx")

sheet = wb.active

print("BU PROGRAM, MAZAK BORU LAZERDE MİNİMAL METARİAL PARAMETRESİ İLE KESİLEN ZAVİYELİ PARÇALARDA MEYDANA GELEN PARÇA UZUNLUĞUNDAKİ KISALMAYI HESAPLAMAK İÇİN, POLİGON PLANLAMA BİRİMİ TARAFINDAN OLUŞTURULMUŞTUR.")
print("LÜTFEN AÇILAN RESİMDEKİ YÖNERGELERE UYARAK İSTENİLEN DEĞERLERİ GİRİNİZ.\nİstenilen Değerler için Ondalıklı Sayı Yazmayınız!\nAçı İçin Minimum Değer 1°, Maksimum Değer 89°. Kalınlık İçin Minimum Değer 2 mm, Maksimum Değer 12 mm'dir.")
def zaviye(aci, kalinlik):
  for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == aci:
      for j in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=j).value == kalinlik:
          return sheet.cell(row=i, column=j).value
  return None

# Resim ekleme kodu
webbrowser.open("gorsel_zaviye.png")

while True:

  aci = input("\nAçıyı Giriniz: ")
  kalinlik = input("Hammadde Kalınlığını Giriniz: ")

  while True:

    yon = int(input("Zaviye Bir Uçta mı, İki Uçta mı? (1 veya 2 Giriniz): "))

    if yon == 1 or yon == 2:
      break
    else:
      print("\n")
      print("Lütfen zaviye sayısına 1 veya 2 olarak cevap veriniz!")
      print("\n")

  sonuc = zaviye(aci, kalinlik)

  if yon == 1:
    sonuc = sonuc

  elif yon == 2 and sonuc is not None:
    sonuc = (sonuc * 2.0)

  if sonuc is None:
    print("\n")
    print("{} Derecelik Açının {} mm Kalınlık Zaviyesi Bulunamadı!".format(aci, kalinlik))
    print("Kalınlık için değer minimum 2 mm, maksimum 12 mm - Açı için değer minimum 1°, maksimum 89° olarak girilebilir! \nLütfen kalınlık ve açı bilgilerini tam sayı olarak yazınız! (Örnek: 60, 33, 21, 3 gibi)")
  else:
    sonuc = round(sonuc,1)
    print("{} Uçtaki Zaviyeli Kesim İçin {} Derecelik Açıda, {} mm Kalınlıkta Oluşan Parça Boyundaki Kısalma: {} MM".format(yon, aci, kalinlik, sonuc))

  while True:

    devam = input("\nTekrar Hesaplama Yapmak İster misiniz? (E/H): ").upper()

    if devam in ("E", "H"):
      break
    else:
      print("Yanlış giriş! Lütfen E veya H olarak cevap veriniz.")

  if devam == "H":
    break
